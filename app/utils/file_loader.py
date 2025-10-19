"""
загружает данные из екселя и json
"""
import os
import json
import xlrd
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from app.exceptions.base import FileExtensionError


class FileLoader():
    def __init__(self, file_name: str):
        self.project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        self.file_path = os.path.join(self.project_root, "uploads", file_name)

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл не найден: {os.path.abspath(self.file_path)}")

    def load_data(self):
        raise NotImplementedError


class ExcelLoader(FileLoader):
    """Загружает данные из Excel-файла"""
    def get_engine(self):
        if self.file_path.lower().endswith('.xlsx'):
            engine = 'openpyxl'
            return engine
        if self.file_path.lower().endswith('.xls'):
            engine = 'xlrd'
            return engine
        raise ValueError("Поддерживаются только .xlsx и .xls")


    def load_data(self, sheet_name: str) -> pd.DataFrame:
        try:
            engine = self.get_engine()
            self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine)
            return self.df
        except:
            raise FileExtensionError("Такой лист в екселе отсутствует")


    def get_sheet_names(self) -> list[str]:
        if self.file_path.lower().endswith('.xlsx'):
            wb = load_workbook(filename=self.file_path, read_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            return sheetnames
        elif self.file_path.lower().endswith('.xls'):
            wb = xlrd.open_workbook(self.file_path)
            sheetnames = wb.sheet_names()
            return sheetnames
        else:
            raise ValueError("Поддерживаются только .xlsx и .xls")


    def clean_dataframe_for_json(self, df: pd.DataFrame) -> list[dict]:
        """
        Заменяем inf на nan
        Приводим все столбцы к object, чтобы на None заменить
        Теперь безопасно заменяем все NaN/NaT на None
        """
        df = df.replace([np.inf, -np.inf], np.nan)
        df = df.astype(object)
        df = df.where(pd.notnull(df), None)
        return df.to_dict(orient="records")


class JsonLoader(FileLoader):
    """Загружает данные из Json-файла"""
    def load_data(self) -> dict:
        with open(self.file_path, "r", encoding="utf-8") as f:
            return json.load(f)


if __name__ == "__main__":
    loader = ExcelLoader("2034d539-a27c-4a99-aace-710f7b9cad75.xls")
    print(loader.get_sheet_names())
    data = loader.load_data("data")
    print(data)

    loader = ExcelLoader("309b91f2-5157-4310-be57-56c220857515.xlsx")
    data = loader.load_data("data")
    print(data)

    loader = JsonLoader("f88bd894-5a23-4cfd-b54d-ead455829630.json")
    data = loader.load_data()
    print(data)
