"""
Задание 2 (Python)
Реализуйте на Python с использованием pandas и подхода ООП инструмент, который:
* загружает данные из Excel-файла
* выполняет тот же расчёт просроченной задолженности, что и в первом задании
* агрегирует данные по РСО (суммирует просрочку)
* сохраняет результат в формате JSON

надо сделать:
апи загрузка файла
апи получить список файлов и листов
апи получить содержимое файла
апи получить отчёт
отчёт проверка расширения (а ещё лучше типа данных)

"""
import os
import json
from openpyxl import load_workbook
import pandas as pd


class FileLoader():
    def __init__(self, file_name: str):
        self.project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        self.file_path = os.path.join(self.project_root, "data", file_name)

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл не найден: {os.path.abspath(self.file_path)}")

    def load_data(self):
        raise NotImplementedError


class ExcelLoader(FileLoader):
    """Загружает данные из Excel-файла"""
    def load_data(self, sheet_name: str):
        self.df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        return self.df

    def get_sheet_names(self) -> list[str]:
        wb = load_workbook(filename=self.file_path, read_only=True)
        return wb.sheetnames


class JsonLoader(FileLoader):
    """Загружает данные из Json-файла"""
    def load_data(self) -> dict:
        with open(self.file_path, "r", encoding="utf-8") as f:
            return json.load(f)


class OverdueCalculator:
    """
        Рассчитывает просроченную задолженность на конец периода по формуле:
        max(0, Дебиторская_на_конец - Кредиторская_на_конец - Начислено_за_период)
        Предполагаются следующие названия столбцов:
        - 'Дебиторская задолженность на конец периода'
        - 'Кредиторская задолженность на конец периода'
        - 'Начислено за период'
    """
    @staticmethod
    def calculate(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()

        required_cols = [
            "Дебиторская задолженность на конец периода",
            "Кредиторская задолженность на конец периода",
            "Начислено за период"
        ]

        for col in required_cols:
            if col not in required_cols:
                raise KeyError(f"Отсутствует обязательный столбец: '{col}'")
            # Заменяем неразрывные пробелы на обычные и конвертируем в float
            df[col] = (
                df[col].astype(str)
                .str.replace("\xa0", " ", regex=False)
                .str.replace(" ", "", regex=False)
                .replace("", "0")
                .astype(float)
            )
        overdue = df[required_cols[0]] - df[required_cols[1]] - df[required_cols[2]]
        df["Просроченная задолженность на конец периода"] = overdue.clip(lower=0)
        return df


class RSOAggregator:
    """Выполняет агрегацию по РСО"""
    @staticmethod
    def aggregate(df: pd.DataFrame) -> dict[str, float]:
        if "РСО" not in df.columns or "Просроченная задолженность на конец периода" not in df.columns:
            raise ValueError("Требуемые столбцы отсутствуют")
        return (
            df.groupby("РСО")["Просроченная задолженность на конец периода"]
            .sum()
            .round(2)
            .to_dict()
        )


class DebtReport:
    def __init__(
        self,
        input_file: str,
        sheet_name: str,
        output_file: str
    ):
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.output_file = output_file

    def generate(self):
        # 1. Загрузка
        loader = ExcelLoader(self.input_file)
        df = loader.load_data(self.sheet_name)

        # 2. Расчёт
        df_with_overdue = OverdueCalculator.calculate(df)

        # 3. Агрегация
        result = RSOAggregator.aggregate(df_with_overdue)

        return result


f1 = ExcelLoader("Тестовое задание.xlsx")
# f2 = JsonLoader("overdue_by_rso.json")
print(f1.get_sheet_names())
# print(f1.load_data("task"))
# print(f2.load_data())

report = DebtReport(
    input_file="Тестовое задание.xlsx",
    sheet_name="raw_data",
    output_file="ttt.json"
)
result = report.generate()
print(result)
